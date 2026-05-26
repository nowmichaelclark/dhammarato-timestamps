#!/usr/bin/env python3
"""
make_xlsx.py
============
Converts data/timestamps.csv into data/timestamps.xlsx with
clickable hyperlinks in the Timestamp Link column.

Run any time after main.py to refresh the Excel file:
    python make_xlsx.py
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERROR] openpyxl not found. Run:  pip install openpyxl")

CSV_FILE = Path("data/timestamps.csv")
OUT_FILE = Path("data/timestamps.xlsx")

HEADERS = [
    "Timestamp Link", "Video Title", "Author",
    "Comment", "Timestamp", "Video ID", "Date Added",
]
LINK_COL = HEADERS.index("Timestamp Link") + 1  # 1-based column index

COLUMN_WIDTHS = [12, 42, 20, 60, 10, 12, 12]


def build_xlsx():
    if not CSV_FILE.exists():
        sys.exit(f"[ERROR] {CSV_FILE} not found — run main.py first.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Timestamps"

    # ── Header row ──────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col, heading in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────────────
    row_font  = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    alt_fill  = PatternFill("solid", start_color="EAF0FB")

    with open(CSV_FILE, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for excel_row, record in enumerate(reader, start=2):
            is_alt = (excel_row % 2 == 0)
            for col, key in enumerate(HEADERS, 1):
                val  = record.get(key, "")
                cell = ws.cell(row=excel_row, column=col, value=val)

                if col == LINK_COL and val.startswith("http"):
                    cell.hyperlink = val
                    cell.value     = "▶ Watch"
                    cell.font      = link_font
                else:
                    cell.font = row_font
                    if is_alt:
                        cell.fill = alt_fill

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=(col == HEADERS.index("Comment") + 1),
                )

    # ── Column widths ─────────────────────────────────────────────────────
    for col, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(OUT_FILE)
    print(f"  Saved {OUT_FILE}  ({excel_row - 1} rows)")


if __name__ == "__main__":
    print("Converting CSV → Excel …")
    build_xlsx()
    print("Done.")
